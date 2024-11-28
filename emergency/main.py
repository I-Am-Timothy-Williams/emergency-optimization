import tkinter as tk
import random
import datetime
import numpy as np  # Import numpy for Poisson distribution
from roomgenerator import *
from patientgenerator import PatientGenerator
import tkinter.messagebox as messagebox  # Add this import at the top of main.py
import openpyxl
import csv
from openpyxl.styles import Alignment
from itertools import product

class RoomPlanner(tk.Tk):
    def __init__(self,batch_mode=False):
        super().__init__()
        self.title("Hospital Room Planner")
        self.geometry("800x600")
        self.rooms = {}
        self.room_widgets = []
        self.grid_size = 4
        self.patient_generator = PatientGenerator()
        self.batch_mode = batch_mode

        # Initialize running totals
        self.total_a_patients = 0
        self.total_b_patients = 0
        self.total_c_patients = 0
        self.total_revenue = 0
        self.total_waiting_cost = 0
        self.total_lwbs_cost = 0
        self.total_harmed_cost = 0
        self.total_staffing_cost = 0
        self.total_costs = 0
        self.average_utilization = 0

        self.a_count = 0
        self.b_count = 0
        self.c_count = 0

        self.iteration_count = 0

        self.input_frame = tk.Frame(self)
        self.input_frame.pack(pady=10)
        self.create_input_fields()

        self.current_time = datetime.datetime.strptime("06:00", "%H:%M")
        # Create a persistent time label
        # self.time_label = tk.Label(self, text=self.format_time_label(), font=("Helvetica", 16))
        # self.time_label.pack(pady=10)

        self.canvas = tk.Canvas(self, width=800, height=600, bg="white")
        self.canvas.pack(fill="both", expand=True)

        self.grid_cells = []  # Keep track of grid cell coordinates
        self.initialize_grid()

        self.initialize_excel("simulation_results.xlsx")

        self.distribution_type = None
        self.distribution_parameters = {}
        self.dragging_patient = None

    def initialize_excel(self, file_name):
        """Create an Excel file with the required headers and two sheets."""
        wb = openpyxl.Workbook()
    
        # Sheet 1: Simulation Results
        sheet1 = wb.active
        sheet1.title = "Simulation Results"
    
        # Add Headers for Sheet 1
        headers = [
            "Hour",
            "# of Busy Exam Rooms",
            None,
            None,
            "# Patients in Waiting Room",
            None,
            None,
            "# Patients Roomed Above Triage",
            "# Patients Left Without Being Seen",
            "# Patients Harmed While Waiting"
        ]
        sub_headers = [
            None,
            "A", "B", "C",
            "A", "B", "C",
            None, None, None
        ]
    
        sheet1.append(headers)
        sheet1.append(sub_headers)
    
        # Merge header cells for Sheet 1
        sheet1.merge_cells("B1:D1")
        sheet1.merge_cells("E1:G1")
        sheet1.merge_cells("H1:H2")
        sheet1.merge_cells("I1:I2")
        sheet1.merge_cells("J1:J2")
    
        # Center align the headers
        for col in range(1, 11):
            sheet1.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center")
            sheet1.cell(row=2, column=col).alignment = Alignment(horizontal="center", vertical="center")
    
        # Sheet 2: Profit and Costs
        sheet2 = wb.create_sheet(title="Profit and Costs")
    
        # Add Headers for Sheet 2
        sheet2_headers = [
            ["ROOM UTILIZATION", "Value", "Calculation", "Score"],
            ["Total patient-hours in A exam rooms", 0, "= (24 x # of A exam rooms)", 0],
            ["Total patient-hours in B exam rooms", 0, "= (24 x # of B exam rooms)", 0],
            ["Total patient-hours in C exam rooms", 0, "= (24 x # of C exam rooms)", 0],
            ["Average overall utilization", 0, "= sum of above 3 values ÷ total rooms", 0],
            [],
            ["REVENUE", "", "", ""],
            ["Total A patients served", 0, "x $1,000", 0],
            ["Total B patients served", 0, "x $600", 0],
            ["Total C patients served", 0, "x $250", 0],
            ["TOTAL REVENUE", "", "", 0],
            [],
            ["COSTS", "", "", ""],
            ["Total A patient-hours in waiting room", 0, "x $250", 0],
            ["Total B patient-hours in waiting room", 0, "x $100", 0],
            ["Total C patient-hours in waiting room", 0, "x $25", 0],
            ["Total patients LWBS", 0, "x $200", 0],
            ["Total patients harmed while waiting", 0, "x $10,000", 0],
            [],
            ["Staffing costs:", "", "", ""],
            ["# of A exam rooms", 0, "x $3,900", 0],
            ["# of B exam rooms", 0, "x $3,000", 0],
            ["# of C exam rooms", 0, "x $1,600", 0],
            ["TOTAL COSTS", "", "", 0],
            [],
            ["OPERATING PROFITS", "", "", 0]
        ]
    
        for row in sheet2_headers:
            sheet2.append(row)
    
        # Format and align cells for Sheet 2
        for row in sheet2.iter_rows(min_row=1, max_row=sheet2.max_row, max_col=4):
            for cell in row:
                if cell.value:  # Apply alignment only to non-empty cells
                    cell.alignment = Alignment(horizontal="center", vertical="center")
    
        # Save the workbook
        wb.save(file_name)
    
    def update_excel(self, file_name, hour, busy_rooms, waiting_patients, roomed_above_triage, left_without_being_seen, harmed_from_neglect):
        """Update only the Simulation Results sheet in the Excel file."""
        wb = openpyxl.load_workbook(file_name)
        
        # Explicitly target the "Simulation Results" sheet
        if "Simulation Results" in wb.sheetnames:
            sheet = wb["Simulation Results"]
        else:
            raise ValueError("Simulation Results sheet not found in the workbook.")
    
        # Prepare the data for the row
        row_data = [
            hour,
            busy_rooms.get("A", 0), busy_rooms.get("B", 0), busy_rooms.get("C", 0),
            waiting_patients.get("A", 0), waiting_patients.get("B", 0), waiting_patients.get("C", 0),
            roomed_above_triage,
            left_without_being_seen,
            harmed_from_neglect
        ]
    
        # Append the row to the "Simulation Results" sheet
        sheet.append(row_data)

        # Save the updated file
        wb.save(file_name)



    def create_input_fields(self):
        tk.Label(self.input_frame, text="Number of A Rooms:").grid(row=0, column=0)
        self.a_count_entry = tk.Entry(self.input_frame)
        self.a_count_entry.grid(row=0, column=1)

        tk.Label(self.input_frame, text="Number of B Rooms:").grid(row=1, column=0)
        self.b_count_entry = tk.Entry(self.input_frame)
        self.b_count_entry.grid(row=1, column=1)

        tk.Label(self.input_frame, text="Number of C Rooms:").grid(row=2, column=0)
        self.c_count_entry = tk.Entry(self.input_frame)
        self.c_count_entry.grid(row=2, column=1)

        self.generate_button = tk.Button(self.input_frame, text="Generate Rooms", command=self.generate_rooms)
        self.generate_button.grid(row=3, columnspan=2)

    def initialize_grid(self):
        """Initialize the grid cells for snapping."""
        cell_width = 800 // self.grid_size
        cell_height = (600 - 100) // self.grid_size  # Reduced height to accommodate the waiting room
        for i in range(self.grid_size):
            for j in range(self.grid_size):
                x = i * cell_width
                y = 100 + j * cell_height  # Start below the waiting room
                self.grid_cells.append((x, y, x + cell_width, y + cell_height))


    def calculate_total_cost_and_rooms(self):
        """Calculate the total cost and total number of rooms."""
        total_cost = 0
        total_rooms = 0

        for room_name, characteristics in self.rooms.items():
            if room_name != "WaitingRoom":  # Exclude the waiting room
                total_cost += characteristics.get("cost_per_day", 0)
                total_rooms += 1

        return total_cost, total_rooms

    def confirm_room_selection(self):
        """Display a confirmation dialog for the total cost and rooms."""
        total_cost, total_rooms = self.calculate_total_cost_and_rooms()

        if total_cost > 42000 or total_rooms > 16:
            if self.batch_mode:
                return False
            messagebox.showerror("Error", f"Invalid room configuration: Cost ({total_cost}) or room count ({total_rooms}) exceeded.")
            return False

        if self.batch_mode:
            return True  # Automatically approve in batch mode


        # Show confirmation and proceed
        proceed = messagebox.askyesno(
            "Confirmation",
            f"Total Cost: {total_cost}\nTotal Rooms: {total_rooms}\n\nDo you want to proceed?"
        )
        return proceed

    def generate_rooms(self):
        """Generate rooms and validate selection before proceeding."""
        a_count = int(self.a_count_entry.get())
        b_count = int(self.b_count_entry.get())
        c_count = int(self.c_count_entry.get())

        self.a_count = a_count
        self.b_count = b_count
        self.c_count = c_count
        room_manager = RoomManager()
        self.rooms = room_manager.generate_rooms(a_count, b_count, c_count)

    # Validate and confirm before proceeding
        if not self.confirm_room_selection():
            return

        self.canvas.delete("all")
        self.room_widgets.clear()
        self.create_waiting_room()
        self.create_grid()
        self.snap_rooms_to_grid()
        if self.batch_mode == False:
            self.show_distribution_options()
        else:
            self.confirm_distribution()


    def create_waiting_room(self):
        """Create the waiting room above the 16-cell grid."""
        self.canvas.create_rectangle(0, 0, 800, 100, fill="lightgray", tags="WaitingRoom")
        self.canvas.create_text(400, 50, text="Waiting Room", font=("Helvetica", 16), tags="WaitingRoom")

    def create_grid(self):
        """Draw the 4x4 grid on the canvas below the waiting room."""
        cell_width = 800 // self.grid_size
        cell_height = (600 - 100) // self.grid_size
        for i in range(self.grid_size + 1):
            self.canvas.create_line(i * cell_width, 100, i * cell_width, 600, fill="gray")
            self.canvas.create_line(0, 100 + i * cell_height, 800, 100 + i * cell_height, fill="gray")

    def snap_rooms_to_grid(self):
        """Snap each room into a grid cell."""
        room_colors = {"A": "red", "B": "blue", "C": "yellow"}
        index = 0

        for room_name, characteristics in self.rooms.items():
            if index >= len(self.grid_cells):
                break  # Ensure we don't exceed the number of grid cells
            
            if room_name == 'WaitingRoom':
                continue
            x1, y1, x2, y2 = self.grid_cells[index]
            room_type = room_name[0]  # Get the room type (A, B, or C)
            color = room_colors.get(room_type, "lightblue")

            # Draw the room as a rectangle in the grid cell
            room_rect = self.canvas.create_rectangle(
                x1, y1, x2, y2, fill=color, tags=room_name
            )
            label = self.canvas.create_text((x1 + x2) // 2, (y1 + y2) // 2, text=room_name)

            self.room_widgets.append((room_rect, label))
            index += 1


    def show_distribution_options(self):
        """Show a drop-down menu for selecting the distribution type and create input fields dynamically."""
        for widget in self.input_frame.winfo_children():
            widget.destroy()

        tk.Label(self.input_frame, text="Select Distribution Type:").pack()
        self.distribution_var = tk.StringVar(self)
        self.distribution_var.set("Poisson")
        distribution_menu = tk.OptionMenu(
            self.input_frame, self.distribution_var, "Poisson", "Uniform", "Normal", command=self.update_distribution_fields
        )
        distribution_menu.pack()
        self.parameter_frame = tk.Frame(self.input_frame)
        self.parameter_frame.pack()

        confirm_button = tk.Button(self.input_frame, text="Confirm", command=self.confirm_distribution)
        confirm_button.pack()

        tk.Label(self.input_frame, text="Select Auto-Solve Option:").pack()
        self.auto_solve_var = tk.StringVar(self)
        self.auto_solve_var.set("None")  # Default option
        auto_solve_menu = tk.OptionMenu(
            self.input_frame, self.auto_solve_var, "None", "Option 1: Exact Room Type", "Option 2: Any Available Room", "Option 3", "Option 4"
        )
        auto_solve_menu.pack()

        self.update_distribution_fields("Poisson")

    def update_distribution_fields(self, distribution_type):
        """Update input fields based on the selected distribution type for patient types A, B, and C."""
        for widget in self.parameter_frame.winfo_children():
            widget.destroy()

        tk.Label(self.parameter_frame, text=f"{distribution_type} Parameters for Type A:").pack()
        self.a_param_entries = self.create_distribution_fields(self.parameter_frame, distribution_type)

        tk.Label(self.parameter_frame, text=f"{distribution_type} Parameters for Type B:").pack()
        self.b_param_entries = self.create_distribution_fields(self.parameter_frame, distribution_type)

        tk.Label(self.parameter_frame, text=f"{distribution_type} Parameters for Type C:").pack()
        self.c_param_entries = self.create_distribution_fields(self.parameter_frame, distribution_type)

    def create_distribution_fields(self, frame, distribution_type):
        """Helper to create parameter fields for a given distribution type."""
        entries = {}
        if distribution_type == "Poisson":
            tk.Label(frame, text="Lambda (λ):").pack()
            entries["lambda"] = tk.Entry(frame)
            entries["lambda"].pack()
        elif distribution_type == "Uniform":
            tk.Label(frame, text="Lower Bound:").pack()
            entries["lower_bound"] = tk.Entry(frame)
            entries["lower_bound"].pack()
            tk.Label(frame, text="Upper Bound:").pack()
            entries["upper_bound"] = tk.Entry(frame)
            entries["upper_bound"].pack()
        elif distribution_type == "Normal":
            tk.Label(frame, text="Mean (μ):").pack()
            entries["mean"] = tk.Entry(frame)
            entries["mean"].pack()
            tk.Label(frame, text="Standard Deviation (σ):").pack()
            entries["std_dev"] = tk.Entry(frame)
            entries["std_dev"].pack()
        return entries

    def confirm_distribution(self):
        """Store the selected distribution and parameters for each patient type, confirm, then start simulation."""
        if self.batch_mode == False:
            self.distribution_type = self.distribution_var.get()
            self.distribution_parameters = {
                "A": self.get_distribution_parameters(self.a_param_entries),
                "B": self.get_distribution_parameters(self.b_param_entries),
                "C": self.get_distribution_parameters(self.c_param_entries),
            }
        else:
            self.distribution_type = 'Poisson'
            # self.auto_solve_var = "Option 1: Exact Room Type"
            # self.auto_solve_var = "Option 2: Any Available Room"
            # self.auto_solve_var = "Option 3"
            self.auto_solve_var = "Option 3"
            self.distribution_parameters = {'A': {'lambda': 0.875}, 'B': {'lambda': 1.58333333333}, 'C': {'lambda': 1.70833333333}}
            # USE IF DISTRIBUTION TYPE IS NORMAL self.distribution_parameters = {'A': {'mean': 0.875,'std_dev': 1}, 'B': {'mean': 1.58333333333,'std_dev': 1}, 'C': {'mean': 1.70833333333,'std_dev': 1}}
            # USE IF DISTRIBUTION TYPE IS UNIFORM self.distribution_parameters = {'A': {'lower_bound': 0,'upper_bound': 2}, 'B': {'lower_bound': 0,'upper_bound': 3}, 'C': {'lower_bound': 0,'upper_bound': 4}}

            print('potato')
            self.start_simulation()
            
            return 
        # Ask for confirmation
        proceed = messagebox.askyesno(
            "Confirm Distribution",
            f"You've selected {self.distribution_type} distribution with parameters:\n"
            f"Type A: {self.distribution_parameters['A']}\n"
            f"Type B: {self.distribution_parameters['B']}\n"
            f"Type C: {self.distribution_parameters['C']}\n\n"
            "Proceed to generate patients?"
        )

        if proceed:
            self.start_simulation()


    def get_distribution_parameters(self, entries):
        """Extract parameters from entry fields."""
        if self.distribution_type == "Poisson":
            return {"lambda": float(entries["lambda"].get())}
        elif self.distribution_type == "Uniform":
            return {
                "lower_bound": float(entries["lower_bound"].get()),
                "upper_bound": float(entries["upper_bound"].get())
            }
        elif self.distribution_type == "Normal":
            return {
                "mean": float(entries["mean"].get()),
                "std_dev": float(entries["std_dev"].get())
            }


    def start_simulation(self):
        """Start the simulation and generate patients for the first hour."""
        for widget in self.input_frame.winfo_children():
            widget.destroy()
        if not hasattr(self, "time_label"):
            self.time_label = tk.Label(self.input_frame, text=self.format_time_label(), font=("Helvetica", 16))
            self.time_label.pack(pady=10)
        if self.batch_mode == True:
            self.time_label = tk.Label(self.input_frame, text='hi', font=("Helvetica", 16))
            self.time_label.pack(pady=10)
        self.time_label.config(text=self.format_time_label())

        # Generate patients for the first hour
        self.patient_widgets = []
        for patient_type in ["A", "B", "C"]:
            num_patients = self.determine_patient_count(patient_type)
            for _ in range(num_patients):
                arrival_time = self.current_time.strftime("%H:%M:%S")
                patient_profile = self.patient_generator.generate_patient_profile(arrival_time, patient_type)
                self.create_patient_widget(patient_profile)

        # Handle auto-solve based on the selected option
        if self.batch_mode == False:
            auto_solve_option = self.auto_solve_var.get()
        else:
            auto_solve_option = self.auto_solve_var
        if auto_solve_option in ["Option 1: Exact Room Type", "Option 2: Any Available Room", "Option 3", "Option 4"]:
            self.fast_forward_simulation(auto_solve_option)
        else:
            # Add a "Confirm Choices" button for manual play
            self.confirm_button = tk.Button(
                self.input_frame,
                text="Confirm Choices",
                font=("Helvetica", 14, "bold"),
                bg="green",
                fg="white",
                command=self.confirm_choices
            )
            self.confirm_button.pack(pady=10)


    def auto_solve_exact_room_type(self):
        """Automatically assign patients to their exact room type."""
        placement_made = True  # Track whether any placements were made this pass

        while placement_made:
            placement_made = False  # Reset for this iteration

            for patient_icon in self.patient_widgets:
                tags = self.canvas.gettags(patient_icon)
                patient_type = next((t for t in tags if t in ["A", "B", "C"]), None)

                # Find the first available room of the exact type
                for room_name, room_data in self.rooms.items():
                    if room_name[0] == patient_type and not room_data["occupied"]:
                        room_rects = self.canvas.find_withtag(room_name)
                        if room_rects:
                            # Assign the patient to the room
                            room_coords = self.canvas.coords(room_rects[0])
                            self.canvas.coords(patient_icon, room_coords[0] + 10, room_coords[1] + 10, room_coords[0] + 30, room_coords[1] + 30)
                            room_data["occupied"] = True  # Mark the room as occupied
                            print(f"Placing patient {patient_type} in room {room_name}")
                            print(f"Tags before placement: {tags}")
                            print(f"Tags after placement: {self.canvas.gettags(patient_icon)}")
                            # Ensure the patient's tags are correct
                            current_tags = set(tags)
                            # Ensure the hours_left tag exists
                            hours_left_tag = next((t for t in tags if t.startswith("hours_left_")), None)
                            if not hours_left_tag:
                                hours_left_tag = "hours_left_1"  # Default hours_left if missing
                                current_tags.add(hours_left_tag)
                            
                            # Reapply the updated tags to the patient
                            self.canvas.itemconfig(patient_icon, tags=tuple(current_tags))

                            # Remove the patient from the waiting room
                            self.patient_widgets.remove(patient_icon)
                            placement_made = True
                            break  # Move to the next patient
            # Exit loop if no placements were made this pass
            if not placement_made:
                break

    def auto_solve_any_available_room(self):
        """Automatically assign patients to any available room they can go into."""
        placement_made = True  # Track whether any placements were made this pass

        while placement_made:
            placement_made = False  # Reset for this iteration

            for patient_icon in self.patient_widgets[:]:  # Use [:] to avoid modifying the list during iteration
                tags = self.canvas.gettags(patient_icon)
                patient_type = next((t for t in tags if t in ["A", "B", "C"]), None)

                # Find the first available room the patient can go into
                for room_name, room_data in self.rooms.items():
                    if not room_data.get("occupied", False):  # Check if the room is available
                        room_type = room_name[0]
                        if (
                            (patient_type == "A" and room_type == "A") or
                            (patient_type == "B" and room_type in ["A", "B"]) or
                            (patient_type == "C" and room_type in ["A", "B", "C"])
                        ):
                            room_rects = self.canvas.find_withtag(room_name)
                            if room_rects:
                                # Assign the patient to the room
                                room_coords = self.canvas.coords(room_rects[0])
                                self.canvas.coords(patient_icon, room_coords[0] + 10, room_coords[1] + 10, room_coords[0] + 30, room_coords[1] + 30)
                                room_data["occupied"] = True  # Mark the room as occupied
                                print(f"Placing patient {patient_type} in room {room_name}")
                                print(f"Tags before placement: {tags}")
                                print(f"Tags after placement: {self.canvas.gettags(patient_icon)}")
                                
                                # Ensure the patient's tags are correct
                                current_tags = set(tags)
                                hours_left_tag = next((t for t in tags if t.startswith("hours_left_")), None)
                                if not hours_left_tag:
                                    hours_left_tag = "hours_left_1"  # Default hours_left if missing
                                    current_tags.add(hours_left_tag)

                                # Reapply the updated tags to the patient
                                self.canvas.itemconfig(patient_icon, tags=tuple(current_tags))

                                # Remove the patient from the waiting room
                                self.patient_widgets.remove(patient_icon)
                                placement_made = True
                                break  # Move to the next patient
            # Exit loop if no placements were made this pass
            if not placement_made:
                break

    def auto_solve_closest_level_room(self):
        """Automatically assign patients to rooms by their level, assigning to the closest higher level if their level is full."""
        placement_made = True  # Track whether any placements were made this pass

        while placement_made:
            placement_made = False  # Reset for this iteration

            for patient_icon in self.patient_widgets:
                tags = self.canvas.gettags(patient_icon)
                patient_level = next((t for t in tags if t in ["A", "B", "C"]), None)

                print(f"Processing patient {patient_icon} with level {patient_level} and tags {tags}")
                if not patient_level:
                    continue  # Skip if patient doesn't have a valid level tag

                # Find an available room, prioritizing by level
                available_room = None
                for room_name, room_data in self.rooms.items():  # Sort ensures consistent level order
                    room_level = room_name[0]
                    print(room_data)
                    if room_level >= patient_level and not room_data["occupied"]:
                        available_room = room_name
                        print(f"Found available room {available_room} for patient {patient_icon}")
                        break

                if available_room:
                    # Assign the patient to the room
                    room_rects = self.canvas.find_withtag(available_room)
                    if room_rects:
                        room_coords = self.canvas.coords(room_rects[0])
                        print(f"Room {available_room} coordinates: {room_coords}")
                        self.canvas.coords(patient_icon, room_coords[0] + 10, room_coords[1] + 10, room_coords[0] + 30, room_coords[1] + 30)
                        self.rooms[available_room]["occupied"] = True  # Mark the room as occupied
                        print(f"Placing patient {patient_level} in room {available_room}")

                        # Ensure the patient's tags are correct
                        current_tags = set(tags)
                        hours_left_tag = next((t for t in tags if t.startswith("hours_left_")), None)
                        if not hours_left_tag:
                            hours_left_tag = "hours_left_1"  # Default hours_left if missing
                            current_tags.add(hours_left_tag)
                        
                        # Reapply the updated tags to the patient
                        self.canvas.itemconfig(patient_icon, tags=tuple(current_tags))

                        # Remove the patient from the waiting room
                        self.patient_widgets.remove(patient_icon)
                        placement_made = True
                        break  # Move to the next patient
            # Exit loop if no placements were made this pass
            if not placement_made:
                print("No placements made in this pass.")
                break

    def assign_patients_with_thresholds(self):
        """
        Assign patients to rooms based on their level:
        - A goes to A rooms only.
        - B goes to B rooms primarily, can go to A rooms if 75% of A rooms are available and no B rooms are free.
        - C goes to C rooms primarily, can go to B rooms if 50% of B rooms are available and no C rooms are free,
        or to A rooms if 50% of A rooms are available and it can't go into a B or C room.
        """
        def count_available_rooms(level):
            """Count the available rooms for a specific level."""
            return sum(1 for room_name, room_data in self.rooms.items() if room_name[0] == level and not room_data["occupied"])

        for patient_icon in self.patient_widgets[:]:  # Use a copy of the list to modify it during iteration
            tags = self.canvas.gettags(patient_icon)
            patient_level = next((t for t in tags if t in ["A", "B", "C"]), None)
            if not patient_level:
                continue  # Skip if the patient doesn't have a valid level

            assigned_room = None

            if patient_level == "A":
                # A patients go into A rooms only
                assigned_room = next(
                    (room_name for room_name, room_data in self.rooms.items() if room_name[0] == "A" and not room_data["occupied"]),
                    None,
                )

            elif patient_level == "B":
                # B patients prioritize B rooms
                assigned_room = next(
                    (room_name for room_name, room_data in self.rooms.items() if room_name[0] == "B" and not room_data["occupied"]),
                    None,
                )
                # Fallback to A rooms if 75% of A rooms are available and no B rooms are free
                if not assigned_room:
                    total_a_rooms = sum(1 for room_name in self.rooms if room_name[0] == "A")
                    available_a_rooms = count_available_rooms("A")
                    if available_a_rooms / total_a_rooms >= 0.75:
                        assigned_room = next(
                            (room_name for room_name, room_data in self.rooms.items() if room_name[0] == "A" and not room_data["occupied"]),
                            None,
                        )

            elif patient_level == "C":
                # C patients prioritize C rooms
                assigned_room = next(
                    (room_name for room_name, room_data in self.rooms.items() if room_name[0] == "C" and not room_data["occupied"]),
                    None,
                )
                # Fallback to B rooms if 50% of B rooms are available and no C rooms are free
                if not assigned_room:
                    total_b_rooms = sum(1 for room_name in self.rooms if room_name[0] == "B")
                    available_b_rooms = count_available_rooms("B")
                    if available_b_rooms / total_b_rooms >= 0.50:
                        assigned_room = next(
                            (room_name for room_name, room_data in self.rooms.items() if room_name[0] == "B" and not room_data["occupied"]),
                            None,
                        )
                # Fallback to A rooms if 50% of A rooms are available and no B or C rooms are free
                if not assigned_room:
                    total_a_rooms = sum(1 for room_name in self.rooms if room_name[0] == "A")
                    available_a_rooms = count_available_rooms("A")
                    if available_a_rooms / total_a_rooms >= 0.50:
                        assigned_room = next(
                            (room_name for room_name, room_data in self.rooms.items() if room_name[0] == "A" and not room_data["occupied"]),
                            None,
                        )

            # If a room is found, assign the patient
            if assigned_room:
                room_rects = self.canvas.find_withtag(assigned_room)
                if room_rects:
                    room_coords = self.canvas.coords(room_rects[0])
                    self.canvas.coords(patient_icon, room_coords[0] + 10, room_coords[1] + 10, room_coords[0] + 30, room_coords[1] + 30)
                    self.rooms[assigned_room]["occupied"] = True
                    print(f"Assigned patient {patient_level} to room {assigned_room}")
                    # Remove the patient from the waiting list
                    self.patient_widgets.remove(patient_icon)


    def confirm_choices(self):
        """Confirm the current hour's choices, increment time, and update patients."""
        # Increment the simulated time
        if self.current_time.strftime("%H:%M") == "05:00":
            if self.batch_mode == False:
                messagebox.showinfo("End of Cycle", "You have reached 5:00 AM. The simulation will restart.")
            self.current_time = datetime.datetime.strptime("06:00", "%H:%M")  # Reset to 6:00 AM
        else:
            self.current_time += datetime.timedelta(hours=1)  # Increment time by one hour

        # Update the time label dynamically
        self.time_label.config(text=self.format_time_label())

        # Track metrics
        patients_to_remove = []  # Track patients to be removed
        healed_patients = []  # Track details of healed patients

        busy_rooms = {"A": 0, "B": 0, "C": 0}
        waiting_patients = {"A": 0, "B": 0, "C": 0}

        # Track waiting room events and roomed above triage
        left_without_being_seen = 0
        harmed_from_neglect = 0
        roomed_above_triage = 0

        # Process all patients in the canvas
        for patient_icon in self.canvas.find_all():
            tags = self.canvas.gettags(patient_icon)

            # Check if the patient is a valid patient icon
            if any(tag.startswith("patient_") for tag in tags):
                patient_type = next((t for t in tags if t in ["A", "B", "C"]), "Unknown")
                hours_left_tag = next((t for t in tags if t.startswith("hours_left_")), None)

                # Check if the patient is in a valid room
                in_valid_room = False
                for room_name, room_data in self.rooms.items():
                    if room_name != "WaitingRoom" and room_data["occupied"]:
                        room_rects = self.canvas.find_withtag(room_name)
                        if room_rects and self.is_patient_in_room(patient_icon, self.canvas.coords(room_rects[0])):
                            in_valid_room = True
                            busy_rooms[room_name[0]] += 1  # Increment room count

                            # Check if patient is roomed above triage
                            room_type = room_name[0]
                            if (
                                (patient_type == "B" and room_type == "A") or
                                (patient_type == "C" and room_type in ["A", "B"])
                            ):
                                roomed_above_triage += 1

                            # Process hours_left for healing
                            if hours_left_tag:
                                hours_left = int(hours_left_tag.split("_")[2]) - 1
                                if hours_left <= 0:
                                    patients_to_remove.append(patient_icon)
                                    healed_patients.append(f"Type: {patient_type}")
                                    if patient_type == "A":
                                        self.total_a_patients += 1
                                    if patient_type == "B":
                                        self.total_b_patients += 1
                                    if patient_type == "C":
                                        self.total_c_patients += 1
                                    room_data["occupied"] = False  # Free the room
                                else:
                                    # Update the tag with the new hours_left
                                    new_tags = tuple(t for t in tags if not t.startswith("hours_left_")) + (f"hours_left_{hours_left}",)
                                    self.canvas.itemconfig(patient_icon, tags=new_tags)
                            break  # No need to check more rooms for this patient

                if not in_valid_room:
                    # Patient is in the waiting room
                    waiting_patients[patient_type] += 1
                    if patient_type in ["B", "C"]:
                        # Roll a 20-sided die for leaving without being seen
                        if random.randint(1, 20) == 20:
                            patients_to_remove.append(patient_icon)
                            left_without_being_seen += 1
                    elif patient_type == "A":
                        # Roll a 20-sided die for harm from neglect
                        if random.randint(1, 20) == 20:
                            patients_to_remove.append(patient_icon)
                            harmed_from_neglect += 1

        # Remove patients from the canvas
        for patient_icon in patients_to_remove:
            self.canvas.delete(patient_icon)
            # Only remove from self.patient_widgets if the patient exists in the list
            if patient_icon in self.patient_widgets:
                self.patient_widgets.remove(patient_icon)

        # Inform the user
        healed_message = (
            f"The time has been updated to: {self.current_time.strftime('%I:%M %p')} - {(self.current_time + datetime.timedelta(hours=1)).strftime('%I:%M %p')}.\n\n"
        )
        if healed_patients:
            healed_message += "Patients healed and removed:\n" + "\n".join(healed_patients) + "\n\n"
        else:
            healed_message += "No patients have been healed this hour.\n\n"

        # Summary of waiting room events and roomed above triage
        healed_message += f"Patients who left without being seen: {left_without_being_seen}\n"
        healed_message += f"Patients harmed from neglect: {harmed_from_neglect}\n"
        healed_message += f"Patients roomed above triage: {roomed_above_triage}"

        # Format the hour
        hour = f"{(self.current_time - datetime.timedelta(hours=1)).strftime('%I:%M %p')} - {self.current_time.strftime('%I:%M %p')}"

        # Update the Excel file
        self.update_excel(
            "simulation_results.xlsx",
            hour,
            busy_rooms,
            waiting_patients,
            roomed_above_triage,
            left_without_being_seen,
            harmed_from_neglect
        )
        if self.batch_mode == False:
            messagebox.showinfo("Time Updated", healed_message)

        # Add new patients for the next hour
        self.add_new_patients()





    def add_new_patients(self):
        """Add new patients to the waiting room for the current hour."""
        # Generate new patients for each type (A, B, C)
        for patient_type in ["A", "B", "C"]:
            num_patients = self.determine_patient_count(patient_type)
            for _ in range(num_patients):
                arrival_time = self.current_time.strftime("%H:%M:%S")
                patient_profile = self.patient_generator.generate_patient_profile(arrival_time, patient_type)
                self.create_patient_widget(patient_profile)



    def determine_patient_count(self, patient_type):
        """Determine the number of patients for a given type based on the distribution."""
        params = self.distribution_parameters[patient_type]
        if self.distribution_type == "Poisson":
            return np.random.poisson(params["lambda"])
        elif self.distribution_type == "Uniform":
            return random.randint(params["lower_bound"], params["upper_bound"])
        elif self.distribution_type == "Normal":
            return max(1, int(random.gauss(params["mean"], params["std_dev"])))  # Ensure at least 1 patient


    def create_patient_widget(self, patient_profile):
        """Create a draggable patient icon and display it in the waiting room."""
        # Define the Waiting Room boundaries (adjust based on your layout)
        waiting_room_x_start = 10
        waiting_room_x_end = 790  # Slightly less than canvas width to stay inside the room
        waiting_room_y_start = 10
        waiting_room_y_end = 90   # Matches the height of the Waiting Room

        # Calculate the spacing for patient icons
        total_patients = len(self.patient_widgets)
        spacing = 30  # Space between patient icons
        x = waiting_room_x_start + (total_patients % ((waiting_room_x_end - waiting_room_x_start) // spacing)) * spacing
        y = waiting_room_y_start + (total_patients // ((waiting_room_x_end - waiting_room_x_start) // spacing)) * spacing

        # Ensure icons don't go outside the Waiting Room
        if y > waiting_room_y_end - 20:  # 20 is the size of the patient icon
            y = waiting_room_y_start  # Wrap to the next row
            x = waiting_room_x_start + (total_patients % ((waiting_room_x_end - waiting_room_x_start) // spacing)) * spacing

        # Determine the color based on the acuity level
        acuity_level = patient_profile["acuity_level"]
        if acuity_level == "High":
            color = "purple"
        elif acuity_level == "Medium":
            color = "pink"
        else:  # Low
            color = "white"

        # Create the patient icon
        patient_icon = self.canvas.create_oval(
            x, y, x + 20, y + 20, fill=color,
            tags=(f"patient_{len(self.patient_widgets)}", patient_profile["patient_type"], f"hours_left_{patient_profile['hours_left']}")
        )

        # Make the patient draggable
        self.canvas.tag_bind(f"patient_{len(self.patient_widgets)}", "<Button-1>", self.on_drag_start)
        self.canvas.tag_bind(f"patient_{len(self.patient_widgets)}", "<B1-Motion>", self.on_drag_move)
        self.canvas.tag_bind(f"patient_{len(self.patient_widgets)}", "<ButtonRelease-1>", self.on_drag_end)

        self.patient_widgets.append(patient_icon)


    def on_drag_start(self, event):
        """Start dragging a patient."""
        self.dragging_patient = self.canvas.find_withtag("current")[0]

    def on_drag_move(self, event):
        """Move the patient with the mouse."""
        if self.dragging_patient:
            self.canvas.coords(self.dragging_patient, event.x - 10, event.y - 10, event.x + 10, event.y + 10)

    def on_drag_end(self, event):
        """Snap the patient to the nearest room or back to the waiting room with type restrictions."""
        if self.dragging_patient:
            # Get the patient type from the tags
            tags = self.canvas.gettags(self.dragging_patient)
            patient_type = next(tag for tag in tags if tag in ["A", "B", "C"])

            # Iterate through grid cells to find a valid placement
            for x1, y1, x2, y2 in self.grid_cells:
                if x1 <= event.x <= x2 and y1 <= event.y <= y2:
                    # Check if this cell contains a room
                    for room_name, characteristics in self.rooms.items():
                        room_rects = self.canvas.find_withtag(room_name)  # Get all canvas items with this tag
                        if room_rects:  # Ensure there's at least one matching canvas item
                            room_rect = room_rects[0]  # Use the first matching item
                            room_coords = self.canvas.coords(room_rect)
                            if (
                                room_coords[0] == x1
                                and room_coords[1] == y1
                                and room_coords[2] == x2
                                and room_coords[3] == y2
                            ):
                                room_type = room_name[0]  # Room type is the first letter of the room name

                                # Check room occupancy
                                if characteristics["occupied"]:
                                    messagebox.showerror(
                                        "Room Occupied",
                                        f"Room {room_name} is already occupied by another patient."
                                    )
                                    self.snap_to_waiting_room()
                                    return

                                # Apply type-based placement restrictions
                                if (
                                    (patient_type == "A" and room_type != "A") or
                                    (patient_type == "B" and room_type not in ["A", "B"])
                                ):
                                    messagebox.showerror(
                                        "Invalid Placement",
                                        f"Patient of type {patient_type} cannot be placed in room type {room_type}."
                                    )
                                    self.snap_to_waiting_room()
                                    return

                                # Mark the room as occupied
                                characteristics["occupied"] = True

                                # Snap the patient to the room
                                self.canvas.coords(self.dragging_patient, x1 + 10, y1 + 10, x1 + 30, y1 + 30)
                                self.canvas.addtag_withtag(room_name, self.dragging_patient)  # Tag the patient with the room name
                                self.dragging_patient = None
                                return

            # If no valid room was found, snap back to the waiting room
            self.snap_to_waiting_room()


    def snap_to_waiting_room(self):
        """Snap the dragged patient back to the waiting room and update room occupancy."""
        # Reset the `occupied` attribute if the patient was tagged with a room
        tags = self.canvas.gettags(self.dragging_patient)
        for tag in tags:
            if tag in self.rooms:
                self.rooms[tag]["occupied"] = False  # Mark the room as unoccupied

        # Move the patient back to the waiting room
        index = self.patient_widgets.index(self.dragging_patient)
        x = 50 + index * 100
        self.canvas.coords(self.dragging_patient, x, 20, x + 20, 40)
        self.dragging_patient = None


    def format_time_label(self):
        """Format the current simulated time as a string."""
        next_hour = self.current_time + datetime.timedelta(hours=1)
        return f"Time: {self.current_time.strftime('%I:%M %p')} - {next_hour.strftime('%I:%M %p')}"

    def is_patient_in_room(self, patient, room_coords):
        """Check if a patient is within the bounds of a room."""
        patient_coords = self.canvas.coords(patient)
        return (
            room_coords[0] <= patient_coords[0] <= room_coords[2]
            and room_coords[1] <= patient_coords[1] <= room_coords[3]
        )
    
    def fast_forward_simulation(self, auto_solve_option):
        """Fast-forward through the simulation, automatically solving each hour."""
        while True:
            # Auto-solve the current hour
            if auto_solve_option == "Option 1: Exact Room Type":
                self.auto_solve_exact_room_type()
            elif auto_solve_option == "Option 2: Any Available Room":
                self.auto_solve_any_available_room()
            elif auto_solve_option == "Option 3":
                self.auto_solve_closest_level_room()
            elif auto_solve_option == "Option 4":
                self.assign_patients_with_thresholds()
            # Ensure `confirm_choices` is called to handle healing and progression
            self.confirm_choices()

            # Check if the simulation has completed the last hour
            if self.current_time.strftime("%H:%M") == "06:00":
                break
        
                # Finalize Sheet 2 with calculations
        self.finalize_sheet_two(
            "simulation_results.xlsx",
            num_a_rooms=self.a_count,
            num_b_rooms=self.b_count,
            num_c_rooms=self.c_count,
            num_a_patients=self.total_a_patients,
            num_b_patients=self.total_b_patients,
            num_c_patients=self.total_c_patients,
        )
        # if self.batch_mode == False:
            # Display a summary message at the end
            # messagebox.showinfo(
            #     "Simulation Complete",
            #     "The simulation has completed. You can review the results in the Excel file."
            # )
        self.clear_patients()
        self.total_a_patients=0
        self.total_b_patients=0
        self.total_c_patients=0
        if self.iteration_count == 0:
            room_combinations = [
    (3,5,7),(3,4,4),(3,5,4),(3,6,3),(3,4,3),(3,5,5),(3,8,3),(3,4,5),(3,5,6),(3,4,8),
    (3,5,3),(3,4,6),(3,6,7),(3,7,4),(3,7,5),(3,6,5),(3,4,9),(3,3,6),(3,3,4),(3,3,3),
    (3,3,5),(3,3,9),(3,3,8),(3,3,10),(5,7,0)
]

            self.run_batch_simulations(room_combinations, num_simulations_per_combination=15)
        
    def clear_patients(self):
        """Clear all patient widgets and reset patient-related data."""
        # Delete all patient icons from the canvas
        for patient_icon in self.patient_widgets:
            self.canvas.delete(patient_icon)
        
        # Reset the list of patient widgets
        self.patient_widgets = []
    # def finalize_sheet_two(self, file_name, num_a_rooms, num_b_rooms, num_c_rooms, num_a_patients, num_b_patients, num_c_patients):
    #     """Finalize the Profit and Costs sheet with calculations after the simulation."""
    #     wb = openpyxl.load_workbook(file_name)
    #     sheet1 = wb["Simulation Results"]
    #     sheet2 = wb["Profit and Costs"]

    #     # Formulas for room utilization
    #     sheet2["B2"] = f"=SUM('Simulation Results'!B3:B26)"  # Total A room hours
    #     sheet2["B3"] = f"=SUM('Simulation Results'!C3:C26)"  # Total B room hours
    #     sheet2["B4"] = f"=SUM('Simulation Results'!D3:D26)"  # Total C room hours
    #     sheet2["B5"] = f"=SUM(B2:B4)"                        # Total room hours
    #     sheet2["D2"] = f"=B2/(24*{num_a_rooms})"             # Utilization for A rooms
    #     sheet2["D3"] = f"=B3/(24*{num_b_rooms})"             # Utilization for B rooms
    #     sheet2["D4"] = f"=B4/(24*{num_c_rooms})"             # Utilization for C rooms
    #     sheet2["D5"] = "=SUM(D2:D4)/3"                       # Average utilization

    #     # Formulas for revenue
    #     sheet2["B8"] = num_a_patients                        # Total A patients served
    #     sheet2["B9"] = num_b_patients                        # Total B patients served
    #     sheet2["B10"] = num_c_patients                      # Total C patients served
    #     sheet2["D11"] = f"=SUM(D8:D10)"                    # Total Revenue
    #     sheet2["D8"] = f"=B8*1000"                           # Revenue from A patients
    #     sheet2["D9"] = f"=B9*600"                            # Revenue from B patients
    #     sheet2["D10"] = f"=B10*250"                          # Revenue from C patients

    #     # Formulas for waiting costs, LWBS, and harm
    #     sheet2["B14"] = f"=SUM('Simulation Results'!E3:E26)"  # Total waiting A patients
    #     sheet2["B15"] = f"=SUM('Simulation Results'!F3:F26)"  # Total waiting B patients
    #     sheet2["B16"] = f"=SUM('Simulation Results'!G3:G26)"  # Total waiting C patients
    #     sheet2["B17"] = f"=SUM('Simulation Results'!I3:I26)"  # Patients who left without being seen
    #     sheet2["B18"] = f"=SUM('Simulation Results'!J3:J26)"  # Patients harmed while waiting
    #     sheet2["D14"] = f"=B14*250"
    #     sheet2["D15"] = f"=B15*100"
    #     sheet2["D16"] = f"=B16*25"
    #     sheet2["D17"] = f"=B17*200"
    #     sheet2["D18"] = f"=B18*10000"

    #     # Formulas for staffing costs
    #     sheet2["B21"] = num_a_rooms
    #     sheet2["B22"] = num_b_rooms
    #     sheet2["B23"] = num_c_rooms
    #     sheet2["D21"] = f"=B21*3900"
    #     sheet2["D22"] = f"=B22*3000"
    #     sheet2["D23"] = f"=B23*1600"
    #     sheet2["D24"] = f"=SUM(D14:D23)"

    #     # Formulas for Operating Profits
    #     sheet2["D26"] = f"=D11-D24"

    #     # Formulas for staffing costs
    #     # Save the workbook
    #     wb.save(file_name)
    #     wb.close()
    def finalize_sheet_two(self, file_name, num_a_rooms, num_b_rooms, num_c_rooms, num_a_patients, num_b_patients, num_c_patients):
        """Finalize the Profit and Costs sheet with calculated values after the simulation."""
        wb = openpyxl.load_workbook(file_name)
        sheet1 = wb["Simulation Results"]
        sheet2 = wb["Profit and Costs"]

        # Calculate room utilization
        total_a_room_hours = sum(sheet1[f"B{row}"].value or 0 for row in range(3, 26))
        total_b_room_hours = sum(sheet1[f"C{row}"].value or 0 for row in range(3, 26))
        total_c_room_hours = sum(sheet1[f"D{row}"].value or 0 for row in range(3, 26))
        total_room_hours = total_a_room_hours + total_b_room_hours + total_c_room_hours

        utilization_a = total_a_room_hours / (24 * num_a_rooms) if num_a_rooms else 0
        utilization_b = total_b_room_hours / (24 * num_b_rooms) if num_b_rooms else 0
        utilization_c = total_c_room_hours / (24 * num_c_rooms) if num_c_rooms else 0
        average_utilization = (utilization_a + utilization_b + utilization_c) / 3

        # Write calculated utilization values
        sheet2["B2"] = total_a_room_hours
        sheet2["B3"] = total_b_room_hours
        sheet2["B4"] = total_c_room_hours
        sheet2["B5"] = total_room_hours
        sheet2["D2"] = utilization_a
        sheet2["D3"] = utilization_b
        sheet2["D4"] = utilization_c
        sheet2["D5"] = average_utilization

        # Calculate revenue
        revenue_a = num_a_patients * 1000
        revenue_b = num_b_patients * 600
        revenue_c = num_c_patients * 250
        total_revenue = revenue_a + revenue_b + revenue_c

        # Write revenue values
        sheet2["B8"] = num_a_patients
        sheet2["B9"] = num_b_patients
        sheet2["B10"] = num_c_patients
        sheet2["D8"] = revenue_a
        sheet2["D9"] = revenue_b
        sheet2["D10"] = revenue_c
        sheet2["D11"] = total_revenue

        # Calculate waiting costs
        total_waiting_a = sum(sheet1[f"E{row}"].value or 0 for row in range(3, 27))
        total_waiting_b = sum(sheet1[f"F{row}"].value or 0 for row in range(3, 27))
        total_waiting_c = sum(sheet1[f"G{row}"].value or 0 for row in range(3, 27))
        patients_left = sum(sheet1[f"I{row}"].value or 0 for row in range(3, 27))
        patients_harmed = sum(sheet1[f"J{row}"].value or 0 for row in range(3, 27))

        cost_waiting_a = total_waiting_a * 250
        cost_waiting_b = total_waiting_b * 100
        cost_waiting_c = total_waiting_c * 25
        cost_left = patients_left * 200
        cost_harmed = patients_harmed * 10000

        # Write waiting costs
        sheet2["B14"] = total_waiting_a
        sheet2["B15"] = total_waiting_b
        sheet2["B16"] = total_waiting_c
        sheet2["B17"] = patients_left
        sheet2["B18"] = patients_harmed
        sheet2["D14"] = cost_waiting_a
        sheet2["D15"] = cost_waiting_b
        sheet2["D16"] = cost_waiting_c
        sheet2["D17"] = cost_left
        sheet2["D18"] = cost_harmed

        # Calculate staffing costs
        staffing_cost_a = num_a_rooms * 3900
        staffing_cost_b = num_b_rooms * 3000
        staffing_cost_c = num_c_rooms * 1600
        total_costs = (
            cost_waiting_a
            + cost_waiting_b
            + cost_waiting_c
            + cost_left
            + cost_harmed
            + staffing_cost_a
            + staffing_cost_b
            + staffing_cost_c
        )

        # Write staffing costs
        sheet2["B21"] = num_a_rooms
        sheet2["B22"] = num_b_rooms
        sheet2["B23"] = num_c_rooms
        sheet2["D21"] = staffing_cost_a
        sheet2["D22"] = staffing_cost_b
        sheet2["D23"] = staffing_cost_c
        sheet2["D24"] = total_costs

        # Calculate operating profits
        operating_profit = total_revenue - total_costs
        sheet2["D26"] = operating_profit

        # Save the workbook
        wb.save(file_name)


    def run_batch_simulations(self, room_combinations, num_simulations_per_combination=2, output_file="simulation_results.csv"):
        """
        Run multiple simulations for each combination of room staffing levels and save results to a CSV file.
        
        :param room_combinations: List of tuples containing room staffing levels (A, B, C).
        :param num_simulations_per_combination: Number of simulations to run per combination.
        :param output_file: File path to save the simulation results (CSV format).
        """
        batch_results = []

        self.iteration_count += 1

        # Open CSV file for writing results
        with open(output_file, mode='w', newline='') as file:
            writer = csv.writer(file)
            
            # Write header row
            writer.writerow(["A", "B", "C", "average_utilization", "average_revenue", "average_cost", "average_profit"])
            
            for combination in room_combinations:
                print(f"Running simulations for room combination: A={combination[0]}, B={combination[1]}, C={combination[2]}")
                
                combination_results = {
                    "combination": combination,
                    "utilization": [],
                    "revenue": [],
                    "costs": [],
                    "profits": []
                }

                for sim in range(num_simulations_per_combination):
                    print(f"  Simulation {sim + 1} of {num_simulations_per_combination}")
                    
                    # Reset state and start a single simulation for the current combination
                    self.reset_simulation_variables(*combination)
                    self.fast_forward_simulation(auto_solve_option="Option 3")

                    # Extract results after finalizing the simulation
                    results = self.extract_results_from_excel("simulation_results.xlsx")
                    combination_results["utilization"].append(results["average_utilization"])
                    combination_results["revenue"].append(results["total_revenue"])
                    combination_results["costs"].append(results["total_cost"])
                    combination_results["profits"].append(results["operating_profit"])

                # Calculate averages for the current combination
                averages = {
                    "combination": combination,
                    "average_utilization": sum(combination_results["utilization"]) / num_simulations_per_combination,
                    "average_revenue": sum(combination_results["revenue"]) / num_simulations_per_combination,
                    "average_cost": sum(combination_results["costs"]) / num_simulations_per_combination,
                    "average_profit": sum(combination_results["profits"]) / num_simulations_per_combination,
                }
                batch_results.append(averages)
                
                # Write the averages to the CSV file
                writer.writerow([
                    combination[0], combination[1], combination[2],
                    averages["average_utilization"], averages["average_revenue"],
                    averages["average_cost"], averages["average_profit"]
                ])

                print(f"Averages for combination {combination}: {averages}")

        print(f"All batch simulations complete. Results saved to {output_file}.")
        return batch_results




    def reset_simulation_variables(self,a_count,b_count,c_count):
        """Reset all variables and UI elements to prepare for a new simulation."""
        # Reinitialize key variables
        self.rooms = {}
        self.room_widgets = []
        self.grid_size = 4
        self.patient_generator = PatientGenerator()
        self.total_a_patients = 0
        self.total_b_patients = 0
        self.total_c_patients = 0
        self.total_revenue = 0
        self.total_waiting_cost = 0
        self.total_lwbs_cost = 0
        self.total_harmed_cost = 0
        self.total_staffing_cost = 0
        self.total_costs = 0
        self.average_utilization = 0
        self.a_count = a_count
        self.b_count = b_count
        self.c_count = c_count
        self.current_time = datetime.datetime.strptime("06:00", "%H:%M")
        self.distribution_type = None
        self.distribution_parameters = {}
        self.dragging_patient = None

        # Recreate UI elements
        self.input_frame.destroy()  # Destroy and recreate input frame
        self.input_frame = tk.Frame(self)
        self.input_frame.pack(pady=10)
        self.create_input_fields_batch()

        # self.canvas.destroy()  # Destroy and recreate canvas
        # self.canvas = tk.Canvas(self, width=800, height=600, bg="white")
        # self.canvas.pack(fill="both", expand=True)
        # self.grid_cells = []  # Reset grid cells
        # self.initialize_grid()

        # Reinitialize Excel file (optional, if you want a fresh file per simulation)
        self.initialize_excel("simulation_results.xlsx")

        


    def extract_results_from_excel(self, file_name):
        """Extract results from the finalized Excel file."""
        wb = openpyxl.load_workbook(file_name)
        sheet = wb["Profit and Costs"]

        # # Open the workbook in Excel
        # wb = xw.Book(file_name)
        # sheet = wb.sheets["Profit and Costs"]

        # Fetch values from the finalized Excel sheet
        average_utilization = sheet["D5"].value  # Assuming this cell holds average utilization
        total_revenue = sheet["D11"].value
        total_cost = sheet["D24"].value
        operating_profit = sheet["D26"].value
        # Close the workbook
        wb.close()
        return {
            "average_utilization": average_utilization,
            "total_revenue": total_revenue,
            "total_cost": total_cost,
            "operating_profit": operating_profit,
        }
    def create_input_fields_batch(self):
        tk.Label(self.input_frame, text="Number of A Rooms:").grid(row=0, column=0)
        self.a_count_entry = tk.Entry(self.input_frame)
        self.a_count_entry.grid(row=0, column=1)

        tk.Label(self.input_frame, text="Number of B Rooms:").grid(row=1, column=0)
        self.b_count_entry = tk.Entry(self.input_frame)
        self.b_count_entry.grid(row=1, column=1)

        tk.Label(self.input_frame, text="Number of C Rooms:").grid(row=2, column=0)
        self.c_count_entry = tk.Entry(self.input_frame)
        self.c_count_entry.grid(row=2, column=1)

        self.generate_rooms_batch()

    def generate_rooms_batch(self):
        """Generate rooms and validate selection before proceeding."""
        room_manager = RoomManager()
        self.rooms = room_manager.generate_rooms(self.a_count, self.b_count, self.c_count)

    # Validate and confirm before proceeding
        if not self.confirm_room_selection():
            return
        # messagebox.showinfo(
        #         "Simulation Complete",
        #         "The simulation has completed. You can review the results in the Excel file."
        #     )
        # self.canvas.delete("all")
        self.room_widgets.clear()
        self.create_waiting_room()
        self.create_grid()
        self.snap_rooms_to_grid()
        if self.batch_mode == False:
            self.show_distribution_options()
        else:
            self.confirm_distribution()

if __name__ == "__main__":
    app = RoomPlanner(batch_mode=True)
    # averages = app.run_batch_simulations(num_simulations=30)
    # print("Simulation Averages:", averages)
    app.mainloop()
    
